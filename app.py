from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify, send_from_directory
from datetime import datetime, date
import sqlite3
import os
from werkzeug.utils import secure_filename
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime
from flask import send_file

# ----------------------------
# Konfigurasi dasar Flask
# ----------------------------
app = Flask(__name__)
app.secret_key = 'supersecretkey123'  # ubah ini dengan string acak rahasia

@app.context_processor
def inject_now():
    return {'datetime': datetime}

# ----------------------------
# Konfigurasi Upload
# ----------------------------
UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'pdf', 'doc', 'docx', 'jpg', 'png'}
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# ----------------------------
# Util DB
# ----------------------------
DB_KLASIF = 'klasifikasi.db'
DB_NOMOR = 'nomor_surat.db'  # menyimpan nomor per klasifikasi

def get_conn(db=DB_KLASIF):
    conn = sqlite3.connect(db)
    conn.row_factory = sqlite3.Row
    return conn

# Ambil anak berdasarkan parent (untuk dropdown)
# Ambil anak berdasarkan parent (untuk dropdown)
def get_children(parent=None):
    conn = get_conn(DB_KLASIF)
    cur = conn.cursor()
    if parent:
        cur.execute("""
            SELECT kode, nama
            FROM klasifikasi_surat
            WHERE parent_kode = ?
            ORDER BY kode
        """, (parent,))
    else:
        cur.execute("""
            SELECT kode, nama
            FROM klasifikasi_surat
            WHERE parent_kode IS NULL OR parent_kode = ''
            ORDER BY kode
        """)
    rows = cur.fetchall()
    conn.close()

    # Konversi ke list
    data = [{"kode": r["kode"], "nama": r["nama"]} for r in rows]

    # Tambahkan pilihan "Lainnya" (kode parent tetap sama)
    if parent:
        data.append({
            "kode": parent,               # nilai tetap parent
            "nama": f"Lainnya ({parent})" # teks tampil
        })

    return data

# ----------------------------
# Inisialisasi tabel nomor
# ----------------------------
def init_nomor_table():
    conn = get_conn(DB_NOMOR)
    c = conn.cursor()
    c.execute('''
        CREATE TABLE IF NOT EXISTS nomor_surat (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            kode_klasifikasi TEXT NOT NULL,
            nomor INTEGER NOT NULL,
            suffix TEXT,
            tahun INTEGER NOT NULL,
            bulan INTEGER NOT NULL,
            tanggal TEXT NOT NULL,
            perihal TEXT
        )
    ''')
    conn.commit()
    conn.close()

init_nomor_table()

# ----------------------------
# Helper: romawi & suffix increment
# ----------------------------
ROMAN = ["I","II","III","IV","V","VI","VII","VIII","IX","X","XI","XII"]

def month_to_roman(m):
    return ROMAN[m-1] if 1 <= m <= 12 else str(m)

def next_suffix(suf):
    """Increment suffix: None -> 'a', 'a'->'b', ..., 'z'->'aa'."""
    if not suf:
        return 'a'
    s = list(suf)
    i = len(s)-1
    while i >= 0:
        if s[i] != 'z':
            s[i] = chr(ord(s[i]) + 1)
            return ''.join(s)
        s[i] = 'a'
        i -= 1
    return 'a' + ''.join(s)

# ----------------------------
# Generate preview nomor tanpa insert DB
# ----------------------------
def preview_number_local(kode_klasifikasi, tanggal_obj):
    today = date.today()
    tahun = tanggal_obj.year
    bulan = tanggal_obj.month
    tanggal_str = tanggal_obj.strftime("%Y-%m-%d")
    bulan_rom = month_to_roman(bulan)

    if tanggal_obj > today:
        raise ValueError("Tanggal surat tidak boleh di masa depan.")

    conn = get_conn(DB_NOMOR)
    c = conn.cursor()
    c.execute("""
        SELECT suffix FROM nomor_surat
        WHERE kode_klasifikasi = ? AND tanggal = ?
        ORDER BY id ASC
    """, (kode_klasifikasi, tanggal_str))
    same_day_rows = c.fetchall()
    conn.close()

    day_of_year = tanggal_obj.timetuple().tm_yday
    suffix = None if not same_day_rows else next_suffix(same_day_rows[-1]["suffix"])
    nomor_surat = f"{kode_klasifikasi}/{day_of_year}{('.' + suffix) if suffix else ''}/{bulan_rom}/{tahun}"
    return nomor_surat, day_of_year, suffix, tahun, bulan, tanggal_str

# ----------------------------
# Routes
# ----------------------------
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']

        conn = get_conn(DB_NOMOR)
        cur = conn.cursor()
        cur.execute("SELECT * FROM admin WHERE username=? AND password=?", (username, password))
        user = cur.fetchone()
        conn.close()

        if user:
            session['admin_logged_in'] = True
            session['username'] = user['username']
            session['role'] = user['role']  # simpan role ke session
            return redirect(url_for('index'))
        else:
            return render_template('login.html', error="Username atau password salah.")
    return render_template('login.html')

@app.route('/logout', methods=['GET', 'POST'])
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route('/')
def index():
    if not session.get('admin_logged_in'):
        flash('Silakan login', 'warning')
        return redirect(url_for('login'))

    layer1 = get_children()
    today = date.today().strftime("%Y-%m-%d")
    return render_template('index.html', layer1=layer1, today=today)

@app.route('/get_children')
def get_children_route():
    parent = request.args.get('parent')
    return jsonify(get_children(parent))

# ----------------------------
# Ambil Nomor Surat (aman dari refresh)
# ----------------------------
@app.route('/ambil_nomor', methods=['POST'])
def ambil_nomor():
    kode_klasifikasi = (
        request.form.get('layer4')
        or request.form.get('layer3')
        or request.form.get('layer2')
        or request.form.get('layer1')
    )
    perihal = request.form.get('perihal', '').strip()
    tanggal_str = request.form.get('tanggal_surat')
    file = request.files.get('file_surat')

    if not kode_klasifikasi or not tanggal_str:
        return "Mohon pilih klasifikasi dan tanggal.", 400
    if not file or file.filename == '':
        return "File surat wajib diunggah.", 400

    try:
        tanggal_obj = datetime.strptime(tanggal_str, "%Y-%m-%d").date()
    except ValueError:
        return "Format tanggal salah (gunakan YYYY-MM-DD).", 400

    try:
        # Buat nomor surat sementara (belum insert DB)
        nomor_surat, day_of_year, suffix, tahun, bulan, tanggal_str = preview_number_local(kode_klasifikasi, tanggal_obj)
        safe_nomor = nomor_surat.replace('/', '-')

        # Simpan file dengan nama sesuai nomor surat
        ekstensi = file.filename.rsplit('.', 1)[1].lower()
        filename = f"{safe_nomor}.{ekstensi}"
        save_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(save_path)

        # Setelah file disimpan, baru insert data ke DB
        conn = get_conn(DB_NOMOR)
        c = conn.cursor()
        c.execute("""
            INSERT INTO nomor_surat (kode_klasifikasi, nomor, suffix, tahun, bulan, tanggal, perihal, nama_file)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, (kode_klasifikasi, day_of_year, suffix, tahun, bulan, tanggal_str, perihal, filename))
        new_id = c.lastrowid
        conn.commit()
        conn.close()

        # Redirect ke hasil nomor
        return redirect(url_for('hasil_nomor', surat_id=new_id))

    except Exception as e:
        return f"Error: {e}", 400

# ----------------------------
# Halaman hasil
# ----------------------------
@app.route('/hasil/<int:surat_id>')
def hasil_nomor(surat_id):
    conn = get_conn(DB_NOMOR)
    cur = conn.cursor()
    cur.execute("SELECT * FROM nomor_surat WHERE id = ?", (surat_id,))
    surat = cur.fetchone()
    conn.close()

    if not surat:
        return "Data surat tidak ditemukan.", 404

    bulan_rom = month_to_roman(surat["bulan"])
    nomor_int = surat["nomor"]
    suffix = surat["suffix"]
    nomor_surat = f"{surat['kode_klasifikasi']}/{nomor_int}{('.' + suffix) if suffix else ''}/{bulan_rom}/{surat['tahun']}"

    return render_template('hasil.html', nomor_surat=nomor_surat)

# ----------------------------
# Riwayat & Reset
# ----------------------------
@app.route('/riwayat')
def riwayat():
    conn = get_conn(DB_NOMOR)
    cur = conn.cursor()
    cur.execute("SELECT * FROM nomor_surat ORDER BY id DESC LIMIT 200")
    rows = cur.fetchall()
    conn.close()
    data = [dict(r) for r in rows]
    return jsonify(data)

@app.route('/konfirmasi_reset', methods=['GET', 'POST'])
def konfirmasi_reset():
    if session.get('role') != 'admin':
        return "❌ Akses ditolak: hanya admin yang boleh mereset nomor surat.", 403

    if request.method == 'POST':
        password_input = request.form['password']
        password_admin = '12345'  # 🔒 ganti dengan password admin sebenarnya

        if password_input == password_admin:
            conn = sqlite3.connect('nomor_surat.db')
            cur = conn.cursor()
            cur.execute("DELETE FROM nomor_surat")
            cur.execute("DELETE FROM sqlite_sequence WHERE name='nomor_surat'")
            conn.commit()
            conn.close()

            flash("✅ Database surat berhasil direset!", "success")
            return redirect(url_for('index'))
        else:
            flash("❌ Password salah. Reset dibatalkan.", "danger")
            return redirect(url_for('konfirmasi_reset'))

    return render_template('konfirmasi_reset.html')

@app.route('/riwayat_surat', methods=['GET'])
def riwayat_surat():
    keyword = request.args.get('keyword', '').strip()
    conn = get_conn(DB_NOMOR)
    cur = conn.cursor()

    if keyword:
        cur.execute("""
            SELECT * FROM nomor_surat
            WHERE kode_klasifikasi LIKE ? OR perihal LIKE ? OR tahun LIKE ?
            ORDER BY tanggal DESC, id DESC
        """, (f"%{keyword}%", f"%{keyword}%", f"%{keyword}%"))
    else:
        cur.execute("""
            SELECT * FROM nomor_surat
            ORDER BY tanggal DESC, id DESC
        """)

    data = cur.fetchall()
    conn.close()

    return render_template('riwayat_surat.html', data=data, keyword=keyword)


@app.route('/export_excel')
def export_excel():
    keyword = request.args.get('keyword', '').strip()
    conn = get_conn(DB_NOMOR)
    cur = conn.cursor()

    if session.get('role') != 'admin':
        return "❌ Akses ditolak: hanya admin yang boleh mengekspor data.", 403

    if keyword:
        cur.execute("""
            SELECT * FROM nomor_surat
            WHERE kode_klasifikasi LIKE ? OR perihal LIKE ? OR tahun LIKE ?
            ORDER BY tanggal DESC, id DESC
        """, (f"%{keyword}%", f"%{keyword}%", f"%{keyword}%"))
    else:
        cur.execute("""
            SELECT * FROM nomor_surat
            ORDER BY tanggal DESC, id DESC
        """)

    data = cur.fetchall()
    conn.close()

    # Buat workbook baru
    wb = Workbook()
    ws = wb.active
    ws.title = "Riwayat Surat"

    # ------------------------
    # 1️⃣ Format Header
    # ------------------------
    headers = ["No", "Kode Klasifikasi", "Nomor Surat", "Perihal", "Tanggal"]
    ws.append(headers)

    header_fill = PatternFill(start_color="007bff", end_color="007bff", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # ------------------------
    # 2️⃣ Isi data surat
    # ------------------------
    for i, row in enumerate(data, start=1):
        nomor_surat = f"{row['kode_klasifikasi']}/{row['nomor']}{('.' + row['suffix']) if row['suffix'] else ''}/{row['bulan']}/{row['tahun']}"
        values = [i, row['kode_klasifikasi'], nomor_surat, row['perihal'], row['tanggal']]
        ws.append(values)

        for j in range(1, 6):
            cell = ws.cell(row=i + 1, column=j)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

    # ------------------------
    # 3️⃣ Penataan kolom
    # ------------------------
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 45
    ws.column_dimensions['E'].width = 15

    # ------------------------
    # 4️⃣ Simpan file sementara dengan nama dinamis
    # ------------------------
    date_str = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    filename = f"riwayat_surat_{date_str}.xlsx"
    filepath = os.path.join("uploads", filename)
    wb.save(filepath)

    return send_file(filepath, as_attachment=True)

# ----------------------------
# Akses file upload
# ----------------------------
@app.route('/uploads/<filename>')
def uploaded_file(filename):
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename)

# ----------------------------
# Jalankan server
# ----------------------------
if __name__ == '__main__':
    app.run(debug=True)
